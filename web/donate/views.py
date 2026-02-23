from collections import defaultdict
from django.shortcuts import render
from .forms import UploadCSVForm
from openpyxl import load_workbook

###
import pandas as pd
from .forms import UploadForm

def upload_view(request):
    result_records = []
    result_columns = []
    total_amount = None

    if request.method == "POST":
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['csv_file']
            df = pd.read_excel(excel_file)

            # 假設欄位順序：日期, 編號名稱, 項目明細, 收支別, 金額
            df.iloc[:,0] = pd.to_datetime(df.iloc[:,0])

            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            group_field = form.cleaned_data['group_field']

            # 篩選日期區間
            mask = (df.iloc[:,0] >= pd.to_datetime(start_date)) & (df.iloc[:,0] <= pd.to_datetime(end_date))
            df_filtered = df.loc[mask]

            # 找出對應欄位索引
            col_map = {'編號名稱':1, '項目明細':2, '收支別':3}
            group_col = df.columns[col_map[group_field]]

            # groupby 金額加總
            result = df_filtered.groupby(group_col)[df.columns[4]].sum().reset_index()

            if not result.empty:
                result_records = result.values.tolist()
                result_columns = result.columns.tolist()
                total_amount = result[df.columns[4]].sum()
    else:
        form = UploadForm()

    return render(request, "donate/uploaddonate.html", {
        "form": form,
        "result_records": result_records,
        "result_columns": result_columns,
        "total_amount": total_amount
    })




def upload_xlsx(request):
    result = None
    if request.method == 'POST':
        form = UploadCSVForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']

            try:
                file.seek(0)
                wb = load_workbook(file, data_only=True)
                ws = wb.active

                totals = defaultdict(float)
                first_row = True
                for row in ws.iter_rows(values_only=True):
                    if row and len(row) >= 2:
                        name_or_id, amount = row[0], row[1]

                        # 跳過標題列
                        if first_row:
                            first_row = False
                            continue

                        # 確保金額是數字
                        try:
                            key = str(name_or_id).strip()  # 第一欄轉成字串
                            totals[key] += float(amount)
                        except (TypeError, ValueError):
                            # 如果金額不是數字就跳過
                            continue

                result = dict(totals)

            except Exception as e:
                result = {"Error": f"檔案讀取失敗: {str(e)}"}
    else:
        form = UploadCSVForm()

    return render(request, 'donate/upload.html', {'form': form, 'result': result})
